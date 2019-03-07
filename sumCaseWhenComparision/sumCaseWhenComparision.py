#!/ms/dist/python/PROJ/core/2.7.3/bin/python

'''
Documentation of Script

# Date - 10-Feb-2017
# Owner - Tejas Shende
# Python Version - 2.7.3
# Purpose - This script will prepare the SUM CASE WHEN queries & execute them, also this will prepare the result along with samples.
'''

import ms.version

#ms.version.addpkg("ibm_db","2.0.7-10.5.0") 
ms.version.addpkg("ibm_db","2.0.4-9.5.5")
ms.version.addpkg("ms.db2","1.0.3")        # specify which ms.db2 module want to use 
ms.version.addpkg("ms.modulecmd", "1.0.4") # we will use ms.modulecmd to module load ibmdb2/client/9.5.5 
import ms.modulecmd
ms.modulecmd.load("ibmdb2/client/9.5.5") # load the module 
ms.modulecmd.load("ibmdb2/client/9.5.5") # load the module 
ms.version.addpkg('numpy', '1.7.1-mkl')
ms.version.addpkg('pandas', '0.13.0')
ms.version.addpkg('dateutil', '1.5')
ms.version.addpkg('xlrd', '0.9.2')
ms.version.addpkg('xlwt', '1.1.2')
ms.version.addpkg('openpyxl', '1.8.5')
ms.version.addpkg('xlsxwriter', '0.7.7')

import ibm_db
import ms.db2
import csv
import time
import openpyxl
import xlsxwriter
import xlrd
import xlwt
import os
import pandas as pd
from pandas import ExcelWriter
from openpyxl import writer
from openpyxl import Workbook
from openpyxl import load_workbook

#Declaring the DataFrames
global qaIsNullProdIsNotNullDF
global qaIsNotNullProdIsNullDF
global qaEqualToProdDF
global qaIsNullProdIsNullDF
global mismatchDF
global dataMismatchDF
global qaIsNullProdNotNullMismatchDF
global qaIsNotNullProdNullMismatchDF
global qaIsNullProdNullMismatchDF

#Declaring the DataFrame's Transpose
global qaIsNullProdIsNotNullDFTr
global qaIsNotNullProdIsNullDFTr
global qaEqualToProdDFTr
global qaIsNullProdIsNullDFTr

global rowCount

#Database Connection
dbcon = ms.db2.connect("NYTD_LCDMart") 
dbcur = dbcon.cursor()

class sumCaseWhenComparision(object):

	try:
		
		dir = os.getcwd()
		
		#Reading the inputs from input excel
		filename = openpyxl.reader.excel.load_workbook(dir + '/sumCaseWhenQueries_Inputs.xlsx')
		sheet = filename.get_sheet_by_name('inputs')
		whereClause = "\n" + sheet.rows[1][1].value
		primaryKey = sheet.rows[1][2].value
		
		#Getting the max rows from Input Sheet
		maxRows = sheet.rows
		
		# Function 1 - This function will dynamically prepare the SUM CASE WHEN Query for scenario - QA IS NULL & PROD IS NOT NULL
		def qaIsNullProdIsNotNull(self):

			print('       Executing the SQL for scenario -: QA IS NULL & PROD IS NOT NULL')
			
			selectClause = """SELECT""" + "\n" + """'QA IS NULL & PROD IS NOT NULL' AS compareType,""" + "\n"
			
			i, strSQL = 1, ''
			for i in range(len(self.maxRows)):
				if not (i==0):
					str = "SUM(CASE WHEN qa." + self.sheet.rows[i][0].value +  " IS NULL AND prod." + self.sheet.rows[i][0].value + " IS NOT NULL THEN 1 ELSE 0 END) as " + self.sheet.rows[i][0].value + "," + "\n"
					strSQL += str
					
			#print(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Writing SQL to text file
			writeFile = open('qaIsNullProdIsNotNull.txt', 'w')
			writeFile.writelines(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Executing the QA IS NULL & PROD IS NOT NULL SQL
			dbcur.execute(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#printing the DB Headers
			num_fields = len(dbcur.description)
			fields = [b[0] for b in dbcur.description]
			
			#Writing the above query output to csv file
			with open('qaIsNullProdIsNotNull.csv', 'w') as result:
				writer = csv.writer(result, dialect = 'excel')
				writer.writerow(fields)
				
				for records in dbcur.fetchall():
					#print(records)
					#if not (records == None):
					writer.writerow(records)
			
			#Creating the DF & reading the data from csv
			self.qaIsNullProdIsNotNullDF = pd.DataFrame()
			self.qaIsNullProdIsNotNullDF = pd.read_csv('qaIsNullProdIsNotNull.csv')
						
			#Transposing the DF
			self.qaIsNullProdIsNotNullDFTr = pd.DataFrame()
			self.qaIsNullProdIsNotNullDFTr = self.qaIsNullProdIsNotNullDF.transpose()
		
		# Function 2 - This function will dynamically prepare the SUM CASE WHEN Query for scenario - QA IS NOT NULL AND PROD IS NULL
		def qaIsNotNullProdIsNull(self):
		
			print('       Executing the SQL for scenario -: QA IS NOT NULL & PROD IS NULL')
			
			selectClause = """SELECT""" + "\n" + """'QA IS NOT NULL & PROD IS NULL' AS compareType,""" + "\n"
			
			i, strSQL = 1, ''
			for i in range(len(self.maxRows)):
				if not (i==0):
					str = "SUM(CASE WHEN qa." + self.sheet.rows[i][0].value +  " IS NOT NULL AND prod." + self.sheet.rows[i][0].value + " IS NULL THEN 1 ELSE 0 END) as " + self.sheet.rows[i][0].value + "," + "\n"
					strSQL += str
					
			#Writing SQL to text file
			writeFile = open('qaIsNotNullProdIsNull.txt', 'w')
			writeFile.writelines(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Executing the QA IS NOT NULL AND PROD IS NULL
			dbcur.execute(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#printing the DB Headers
			num_fields = len(dbcur.description)
			fields = [b[0] for b in dbcur.description]
			
			#Writing the above query output to csv file
			with open('qaIsNotNullProdIsNull.csv', 'w') as result:
				writer = csv.writer(result, dialect = 'excel')
				writer.writerow(fields)
				for records in dbcur.fetchall():
					writer.writerow(records)
			
			#Creating the DF & reading the data from csv
			self.qaIsNotNullProdIsNullDF = pd.DataFrame()
			self.qaIsNotNullProdIsNullDF = pd.read_csv('qaIsNotNullProdIsNull.csv')
						
			#Transposing the DF
			self.qaIsNotNullProdIsNullDFTr = pd.DataFrame()
			self.qaIsNotNullProdIsNullDFTr = self.qaIsNotNullProdIsNullDF.transpose()
		
		# Function 3 - This function will dynamically prepare the SUM CASE WHEN Query for scenario - QA EQUAL TO PROD
		def qaEqualToProd(self):
			
			print('       Executing the SQL for scenario -: QA EQUAL TO PROD')
			
			selectClause = """SELECT""" + "\n" + """'QA = PROD' AS compareType,""" + "\n"
			
			i, strSQL = 1, ''
			for i in range(len(self.maxRows)):
				if not (i==0):
					str = "SUM(CASE WHEN qa." + self.sheet.rows[i][0].value +  " = prod." + self.sheet.rows[i][0].value + " THEN 1 ELSE 0 END) as " + self.sheet.rows[i][0].value + "," + "\n"
					strSQL += str
					
			#print(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Writing SQL to text file
			writeFile = open('qaEqualToProd.txt', 'w')
			writeFile.writelines(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Executing the QA EQUAL TO PROD SQL
			dbcur.execute(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#printing the DB Headers
			num_fields = len(dbcur.description)
			fields = [b[0] for b in dbcur.description]
			
			#Writing the above query output to csv file
			with open('qaEqualToProd.csv', 'w') as result:
				writer = csv.writer(result, dialect = 'excel')
				writer.writerow(fields)
				for records in dbcur.fetchall():
					writer.writerow(records)
			
			#Creating the DF & reading the data from csv
			self.qaEqualToProdDF = pd.DataFrame()
			self.qaEqualToProdDF = pd.read_csv('qaEqualToProd.csv')
						
			#Transposing the DF
			self.qaEqualToProdDFTr = pd.DataFrame()
			self.qaEqualToProdDFTr = self.qaEqualToProdDF.transpose()

		# Function 4 - This function will dynamically prepare the SUM CASE WHEN Query for scenario - QA IS NULL AND PROD IS NULL
		def qaIsNullProdIsNull(self):
		
			print('       Executing the SQL for scenario -: QA IS NULL & PROD IS NULL')
			
			selectClause = """SELECT""" + "\n" + """'QA IS NULL & PROD IS NULL' AS compareType,""" + "\n"
			
			i, strSQL = 1, ''
			for i in range(len(self.maxRows)):
				if not (i==0):
					str = "SUM(CASE WHEN qa." + self.sheet.rows[i][0].value +  " IS NULL AND prod." + self.sheet.rows[i][0].value + " IS NULL THEN 1 ELSE 0 END) as " + self.sheet.rows[i][0].value + "," + "\n"
					strSQL += str
					
			#print(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Writing SQL to text file
			writeFile = open('qaIsNullProdIsNull.txt', 'w')
			writeFile.writelines(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#Executing the QA IS NULL & PROD IS NOT NULL SQL
			dbcur.execute(selectClause + strSQL[0:len(strSQL)-2] + self.whereClause)
			
			#printing the DB Headers
			num_fields = len(dbcur.description)
			fields = [b[0] for b in dbcur.description]
			
			#Writing the above query output to csv file
			with open('qaIsNullProdIsNull.csv', 'w') as result:
				writer = csv.writer(result, dialect = 'excel')
				writer.writerow(fields)
				for records in dbcur.fetchall():
					writer.writerow(records)
			
			#Creating the DF & reading the data from csv
			self.qaIsNullProdIsNullDF = pd.DataFrame()
			self.qaIsNullProdIsNullDF = pd.read_csv('qaIsNullProdIsNull.csv')

			#Transposing the DF
			self.qaIsNullProdIsNullDFTr = pd.DataFrame()
			self.qaIsNullProdIsNullDFTr = self.qaIsNullProdIsNullDF.transpose()
			
			#Concatenating all the DataFrames to one Dataframe.
			allDF = pd.DataFrame()
			frames = [self.qaIsNullProdIsNotNullDF, self.qaIsNotNullProdIsNullDF, self.qaEqualToProdDF, self.qaIsNullProdIsNullDF]
			allDF = pd.concat(frames)
			
			#Writing the concatenated DataFrame to csv
			#allDF.transpose().to_csv('sumCaseWhenComparision_output.csv')
						
			#xlWriter = ExcelWriter('sumCaseWhenComparision_Output.xlsx')
			xlWriter = ExcelWriter('sumCaseWhenComparision_Output.xlsx')
			allDF.transpose().to_excel(xlWriter, 'sumCaseWhenComparision')
			
		#This function will find out the count of matching records between QA & PROD
		def getMatchingRowCount(self):
			selectClause = 'SELECT COUNT(*) '
		
			dbcur.execute(selectClause + self.whereClause)
			
			for records in dbcur.fetchall():
				self.rowCount = records[0]
		
		#This function will SUM up the values for Count Mismatch / Data Mismatch
		def sumUpValues(self):
			dir = os.getcwd()
			print('Summing up the values for count mismatch & data mismatch')
			
			wb = Workbook()
			wb = load_workbook(dir + '/sumCaseWhenComparision_Output.xlsx')
			ws = wb.get_sheet_by_name('sumCaseWhenComparision')
			
			ws["F2"].value = "MISMATCH_COUNT"
			
			#i, j, k, rw, col = 2, 5, 0, 2, 3
			i, j, k, rw, col = 2, 5, 0, 2, 1
			
			while(k < len(ws.rows)-2):
				#Segregating the excel formula in to 3 steps
				
				#Formula = TotalMatchingRows - SUM(B3, C3, D3, E3)
				
				#Step - 1
				#var1 pointing to B3 i.e. column is QA IS NULL & PROD IS NOT NULL
				#var2 pointing to C3 i.e. column is QA IS NOT NULL & PROD IS NULL
				#var3 pointing to D3 i.e. column is QA=PROD
				#var4 pointing to E3 i.e. column is QA IS NULL & PROD IS NULL
				
				rowCnt = str(self.rowCount)
				var1 = str(ws.cell(row=rw, column=col).value)
				var2 = str(ws.cell(row=rw, column=col+1).value)
				var3 = str(ws.cell(row=rw, column=col+2).value)
				var4 = str(ws.cell(row=rw, column=col+3).value)
				
				#Step - 2
				#Summing up the above 4 variables value into sumVar variable
				sumVar =  int(var1) + int(var2) + int(var3) + int(var4)
				
				#Step - 3
				#Subtracting the sumVar from rowCnt & Setting the output in MISMATCH_COUNT column
				ws.cell(row=i, column=j).value = int(rowCnt) - sumVar
				
				i+=1
				rw+=1
				k+=1
			
			#Saving & closing the output excel sheet
			wb.save("sumCaseWhenComparision_Output.xlsx")
			wb = None
			ws = None
			
			#Reading the above Excel sheet data into DataFrame
			opxlsFile = pd.ExcelFile('sumCaseWhenComparision_Output.xlsx')
			self.mismatchDF = pd.DataFrame()
			self.mismatchDF = opxlsFile.parse('sumCaseWhenComparision')
			
			# filename = openpyxl.reader.excel.load_workbook(dir + '/sumCaseWhenComparision_Output.xlsx')
			# sheet = filename.get_sheet_by_name('sumCaseWhenComparision')
			# maxRows = sheet.rows
			# sheet.rows[0][0].value = "countMismatch"
			# filename.save('sumCaseWhenComparision_output.xlsx')
			
		#This function will get the mismatch samples
		def getMismatchSamples(self):
			
			wb = Workbook()
			wb = load_workbook('sumCaseWhenComparision_Output.xlsx')
			ws = wb.get_sheet_by_name('sumCaseWhenComparision')
			
			print('Querying the Database for mismatch samples...')
			
			#Getting the mismatch sample for scenario - Data Mismatch  --> #Scenario 1
			
			i, j, k, rw, col = 2, 5, 0, 2, 0
			# i, j =  Pointing to MISMATCH_COLUMNS row (F3)
			# k = Just a counter variable
			# rw, col =  Pointing to Database column Names (A3)
			
			print("       Writing the Data Mismatch records to file")
			
			while(k < len(ws.rows)-2):

				if not (int(ws.cell(row=i, column=j).value)) == 0:
					
					selectClause = "SELECT DISTINCT " + self.primaryKey + ", qa." + ws.cell(row = rw, column = col).value + " as QA_" + ws.cell(row = rw, column = col).value + ", prod." + ws.cell(row = rw, column = col).value + " as PROD_" + ws.cell(row = rw, column = col).value
					
					#print(selectClause + self.whereClause + "AND\nqa." + ws.cell(row = rw, column = col).value + " <> prod." + ws.cell(row = rw, column = col).value + "\nFETCH FIRST 10 ROWS ONLY")

					dbcur.execute(selectClause + self.whereClause + "\n" + "AND qa." + ws.cell(row = rw, column = col).value + " <> prod." + ws.cell(row = rw, column = col).value + "\nFETCH FIRST 10 ROWS ONLY")
					
					## getting the DB Headers
					num_fields = len(dbcur.description)
					fields = [b[0] for b in dbcur.description]
					
					## writing temp mismatch.csv file
					with open('dataMismatch.csv', 'a') as result:
						writer = csv.writer(result, dialect = 'excel')
						writer.writerow(fields)
					
						for records in dbcur.fetchall():
							writer.writerow(records)

					#Reading the data from DataMismatch.csv
					self.dataMismatchDF = pd.DataFrame()
					self.dataMismatchDF = pd.read_csv('dataMismatch.csv')

				k+=1
				i+=1
				rw+=1
			
			#Getting the mismatch samples for scenario - QA IS NULL & PROD IS NOT NULL --> #Scenario 2
			
			i, j, k, rw, col = 2, 1, 0, 2, 0
			# i, j = Pointing to QA IS NULL & PROD IS NOT NULL column (B3)
			# k = Just a counter variable
			# rw, col = Pointing to Database column Names (A3)			
			
			print("       Writing QA IS NULL & PROD IS NOT NULL mismatch records to file")
			
			while(k < len(ws.rows)-2):
				
				#if not(ws.cell(row=i, column=j).value == 0):
				selectClause = "SELECT DISTINCT " + self.primaryKey + ", qa." + ws.cell(row = rw, column = col).value + " as QA_" + ws.cell(row = rw, column = col).value + ", prod." + ws.cell(row = rw, column = col).value + " as PROD_" + ws.cell(row = rw, column = col).value
				
				#print(selectClause + self.whereClause + "\nqa." + ws.cell(row = rw, column = col).value + " IS NULL AND prod." + ws.cell(row = rw, column = col).value + " IS NOT NULL" + "\nFETCH FIRST 10 ROWS ONLY")
				
				dbcur.execute(selectClause + self.whereClause + "\n" + "AND qa." + ws.cell(row = rw, column = col).value + " IS NULL AND prod." + ws.cell(row = rw, column = col).value + " IS NOT NULL" + "\nFETCH FIRST 10 ROWS ONLY")
				
				## printing the DB Headers
				num_fields = len(dbcur.description)
				fields = [b[0] for b in dbcur.description]
				
				## writing temp mismatch.csv file
				with open('qaIsNullProdNotNull.csv', 'a') as result:
					writer = csv.writer(result, dialect = 'excel')
					writer.writerow(fields)
				
					for records in dbcur.fetchall():
						writer.writerow(records)
			
				#Reading the data from qaIsNullProdNotNull.csv
				self.qaIsNullProdNotNullMismatchDF = pd.DataFrame()
				self.qaIsNullProdNotNullMismatchDF = pd.read_csv('qaIsNullProdNotNull.csv')
					
				k+=1
				i+=1
				rw+=1
						
			#Getting the mismatch samples for scenario - QA IS NOT NULL & PROD IS NULL --> #Scenario 3
			
			i, j, k, rw, col = 2, 2, 0, 2, 0
			# i, j = Pointing to QA IS NOT NULL & PROD IS NULL column (C3)
			# k = Just a counter variable
			# rw, col = Pointing to Database column Names (A3)
			
			print("       Writing QA IS NOT NULL & PROD IS NULL mismatch records to file")
			
			while(k < len(ws.rows)-2):
				
				#if not (ws.cell(row=i, column=j).value == 0):
				selectClause = "SELECT DISTINCT " + self.primaryKey + ", qa." + ws.cell(row = rw, column = col).value + " as QA_" + ws.cell(row = rw, column = col).value + ", prod." + ws.cell(row = rw, column = col).value + " as PROD_" + ws.cell(row = rw, column = col).value
				
				#print(selectClause + self.whereClause + "\nqa." + ws.cell(row = rw, column = col).value + " IS NOT NULL AND prod." + ws.cell(row = rw, column = col).value + " IS NOT NULL" + "\nFETCH FIRST 10 ROWS ONLY")
				
				dbcur.execute(selectClause + self.whereClause + "\n" + "AND qa." + ws.cell(row = rw, column = col).value + " IS NOT NULL AND prod." + ws.cell(row = rw, column = col).value + " IS NULL" + "\nFETCH FIRST 10 ROWS ONLY")
				
				## getting the DB Headers
				num_fields = len(dbcur.description)
				fields = [b[0] for b in dbcur.description]
				
				## writing temp mismatch.csv file.
				with open('qaIsNotNullProdIsNull_DM.csv', 'a') as result:
					writer = csv.writer(result, dialect = 'excel')
					writer.writerow(fields)
				
					for records in dbcur.fetchall():
						writer.writerow(records)
						#print(records)
						
				#Reading the data from qaIsNotNullProdIsNull_DM.csv
				self.qaIsNotNullProdNullMismatchDF = pd.DataFrame()
				self.qaIsNotNullProdNullMismatchDF = pd.read_csv('qaIsNotNullProdIsNull_DM.csv')
					
				k+=1
				i+=1
				rw+=1
				
			#Getting the mismatch samples for scenario - QA IS NULL & PROD NULL --> #Scenario 4
			
			i, j, k, rw, col = 2, 4, 0, 2, 0
			# i, j =  Pointing to QA IS NULL & PROD IS NULL column (E3)
			# k = Just a counter variable
			# rw, col =  Pointing to Database column Names (A3)
			
			print("       Writing QA IS NULL & PROD IS NULL mismatch records to file")
			
			while(k < len(ws.rows)-2):
				
				if not(ws.cell(row=i, column=j).value == 0):					
					selectClause = "SELECT DISTINCT " + self.primaryKey + ", qa." + ws.cell(row = rw, column = col).value + " as QA_" + ws.cell(row = rw, column = col).value + ", prod." + ws.cell(row = rw, column = col).value + " as PROD_" + ws.cell(row = rw, column = col).value
					
					#print(selectClause + self.whereClause + "\nqa." + ws.cell(row = rw, column = col).value + " IS NOT NULL AND prod." + ws.cell(row = rw, column = col).value + " IS NOT NULL" + "\nFETCH FIRST 10 ROWS ONLY")
					
					dbcur.execute(selectClause + self.whereClause + "\n" + "AND qa." + ws.cell(row = rw, column = col).value + " IS NULL AND prod." + ws.cell(row = rw, column = col).value + " IS NULL" + "\nFETCH FIRST 10 ROWS ONLY")
					
					## getting the DB Headers
					num_fields = len(dbcur.description)
					fields = [b[0] for b in dbcur.description]
					
					## writing temp mismatch.csv file.
					with open('qaIsNullProdIsNull_DM.csv', 'a') as result:
						writer = csv.writer(result, dialect = 'excel')
						writer.writerow(fields)
					
						for records in dbcur.fetchall():
							writer.writerow(records)
					
					#Reading the data from qaIsNullProdIsNull_DM.csv
					self.qaIsNullProdNullMismatchDF = pd.DataFrame()
					self.qaIsNullProdNullMismatchDF = pd.read_csv('qaIsNullProdIsNull_DM.csv')
					
				k+=1
				i+=1
				rw+=1
				
			#Saving & closing the output excel sheet
			wb.save("sumCaseWhenComparision_Output.xlsx")
			wb = None
			ws = None
			#result.close()
			
		#This function will generated the final report
		def generateReport(self):
			
			print('Generating final report...')
			
			xlWriter = ExcelWriter('caseWhenThenOutput.xlsx')
			
			if (self.mismatchDF.empty):
				print("There are NO Mismatch found...")
			else:
				self.mismatchDF.to_excel(xlWriter, 'report')
			
			if (self.dataMismatchDF.empty):
				print("There are NO Mismatches found for scenario - Data Mismatches")
			else:
				self.dataMismatchDF.to_excel(xlWriter, 'dataMismatch')
				
			if (self.qaIsNullProdNotNullMismatchDF.empty):
				print("There are NO Mismatches found for scenario - QA Is Null & PROD Is Not Null")
			else:
				self.qaIsNullProdNotNullMismatchDF.to_excel(xlWriter, 'QAIsNullProdIsNotNull')
			
			if(self.qaIsNotNullProdNullMismatchDF.empty):
				print("There are NO Mismatches found for scenario - QA Is Not Null & PROD Is Null")
			else:
				self.qaIsNotNullProdNullMismatchDF.to_excel(xlWriter, 'QAIsNotNullProdIsNull')
			
			if(self.qaIsNullProdNullMismatchDF.empty):
				print("There are NO Mismatches found for scenario - QA Is Null & PROD Is Null")
			else:
				self.qaIsNullProdNullMismatchDF.to_excel(xlWriter, 'QAIsNullProdIsNull')
			
			time.sleep(3)
			xlWriter.save()
			time.sleep(5)
			xlWriter.close()

		#This function will remove all temporary files
		def removeTempFile(self):
			dir = os.getcwd()
			files = os.listdir(dir)
			
			print('Removing all Temporay files...')
			
			if 'qaIsNullProdIsNotNull.csv' in files:
				os.remove('qaIsNullProdIsNotNull.csv')
				
			if 'qaIsNotNullProdIsNull.csv' in files:
				os.remove('qaIsNotNullProdIsNull.csv')
				
			if 'qaEqualToProd.csv' in files:
				os.remove('qaEqualToProd.csv')
				
			if 'qaIsNullProdIsNull.csv' in files:
				os.remove('qaIsNullProdIsNull.csv')
			
			if 'qaIsNotNullProdIsNull_DM.csv' in files:
				os.remove('qaIsNotNullProdIsNull_DM.csv')
			
			if 'dataMismatch.csv' in files:
				os.remove('dataMismatch.csv')
			
			if 'qaIsNullProdNotNull.csv' in files:
				os.remove('qaIsNullProdNotNull.csv')
			
			if 'sumCaseWhenComparision_Output.xlsx' in files:
				os.remove('sumCaseWhenComparision_Output.xlsx')
				
			if 'qaIsNullProdIsNull_DM.csv' in files:
				os.remove('qaIsNullProdIsNull_DM.csv')

		##Closing the Database Connection
		def closeDBCon(self):
			print("Closing all open Database Connection(s)...")
			dbcon.close()
			
	except Exception as e:
		print(e)
		dbcon.close()
