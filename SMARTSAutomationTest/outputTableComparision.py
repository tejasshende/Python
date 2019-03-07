#!/ms/dist/python/PROJ/core/2.7.3/bin/python

'''
Documentation of Script

# Date - 20-July-2016
# Owner - Tejas Shende
# Python Version - 2.6.3
# Purpose - This script will compare the SMARTS ouput tables & will prepaer the output with mismatching records.

'''

import ms.version

ms.version.addpkg("ibm_db","2.0.4-9.5.5") 
ms.version.addpkg("ms.db2","1.0.3")        # specify which ms.db2 module want to use 
ms.version.addpkg("ms.modulecmd", "1.0.4") # we will use ms.modulecmd to module load ibmdb2/client/9.5.5 
import ms.modulecmd
ms.modulecmd.load("ibmdb2/client/9.5.5") # load the module 
ms.version.addpkg('numpy', '1.7.1-mkl')
ms.version.addpkg('pandas', '0.13.0')
ms.version.addpkg('dateutil', '1.5')
ms.version.addpkg('xlrd', '0.9.2')
ms.version.addpkg('openpyxl', '1.8.5')
ms.version.addpkg('xlsxwriter', '0.7.7')
ms.version.addpkg('prettytable', '0.7.2')

import ibm_db
import ms.db2
import csv
import openpyxl
import time
import xlsxwriter
import os
import pandas as pd
from prettytable import from_csv
from pandas import ExcelWriter
from openpyxl import writer
from openpyxl import Workbook
from openpyxl import load_workbook

dbcon = ms.db2.connect("NYTD_LCDMart") 
dbcur = dbcon.cursor()
baseDir = os.getcwd()

#Declaring the DataFrames which will be used for Transposing
global enterDFTr
global amendDFTr
global deletDFTr
global tradeDFTr
global mismatchDF

## Setting the options for pandas module.
pd.set_option('display.height', 1000)
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

class outputTableCompare(object):

	## This function will verify the ENTER table
	def verifyEnterTable(self, region, busDt):
		dbcur.execute(
					  '''
					  
						SELECT
						'QA = NULL PROD NOT NULL' AS compareType,
						SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NOT NULL THEN 1 ELSE 0 END) as exchange_id,
						SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NOT NULL THEN 1 ELSE 0 END) as message_type,
						SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NOT NULL THEN 1 ELSE 0 END) as date1,
						SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NOT NULL THEN 1 ELSE 0 END) as time1,
						SUM(CASE WHEN a.order_ID IS NULL AND b.order_ID IS NOT NULL THEN 1 ELSE 0 END) as order_ID,
						SUM(CASE WHEN a.market_side IS NULL AND b.market_side IS NOT NULL THEN 1 ELSE 0 END) as market_side,
						SUM(CASE WHEN a.security IS NULL AND b.security IS NOT NULL THEN 1 ELSE 0 END) as security,
						SUM(CASE WHEN round(a.price,4) IS NULL AND round(b.price,4) IS NOT NULL THEN 1 ELSE 0 END) as price,
						SUM(CASE WHEN round(a.quantity,4) IS NULL AND round(b.quantity,4) IS NOT NULL THEN 1 ELSE 0 END) as quantity,
						SUM(CASE WHEN a.undisclosed_vol IS NULL AND b.undisclosed_vol IS NOT NULL THEN 1 ELSE 0 END) as undisclosed_vol,
						SUM(CASE WHEN a.value1 IS NULL AND b.value1 IS NOT NULL THEN 1 ELSE 0 END) as value1,
						SUM(CASE WHEN trim(a.order_flags) IS NULL AND trim(b.order_flags) IS NOT NULL THEN 1 ELSE 0 END) as order_flags,
						SUM(CASE WHEN a.account IS NULL AND b.account IS NOT NULL THEN 1 ELSE 0 END) as account,
						SUM(CASE WHEN a.accountType IS NULL AND b.accountType IS NOT NULL THEN 1 ELSE 0 END) as accountType,
						SUM(CASE WHEN a.broker IS NULL AND b.broker IS NOT NULL THEN 1 ELSE 0 END) as broker,
						SUM(CASE WHEN a.trader IS NULL AND b.trader IS NOT NULL THEN 1 ELSE 0 END) as trader,
						SUM(CASE WHEN a.order_type IS NULL AND b.order_type IS NOT NULL THEN 1 ELSE 0 END) as order_type,
						SUM(CASE WHEN a.parent_order_id IS NULL AND b.parent_order_id IS NOT NULL THEN 1 ELSE 0 END) as parent_order_id,
						SUM(CASE WHEN a.execution_instrcutions IS NULL AND b.execution_instrcutions IS NOT NULL THEN 1 ELSE 0 END) as execution_instrcutions,
						SUM(CASE WHEN a.order_receive_date IS NULL AND b.order_receive_date IS NOT NULL THEN 1 ELSE 0 END) as order_receive_date,
						SUM(CASE WHEN a.order_receive_time IS NULL AND b.order_receive_time IS NOT NULL THEN 1 ELSE 0 END) as order_receive_time,
						SUM(CASE WHEN a.other IS NULL AND b.other IS NOT NULL THEN 1 ELSE 0 END) as other

						FROM LC.SMARTS_ENTER_QA_''' + region + '_' + busDt + ''' a
						INNER JOIN LC.SMARTS_ENTER_PROD_''' + region + '_' + busDt + ''' b
						ON a.order_ID = b.order_ID
					  
						UNION ALL
						
						SELECT
						'QA NOT NULL PROD = NULL' AS compareType,
						SUM(CASE WHEN a.exchange_id IS NOT NULL AND b.exchange_id IS  NULL THEN 1 ELSE 0 END) as exchange_id,
						SUM(CASE WHEN a.message_type IS NOT NULL AND b.message_type IS  NULL THEN 1 ELSE 0 END) as message_type,
						SUM(CASE WHEN a.date1 IS NOT NULL AND b.date1 IS  NULL THEN 1 ELSE 0 END) as date1,
						SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NOT NULL AND replace(substr(b.time1,1,8),'.',':') IS  NULL THEN 1 ELSE 0 END) as time1,
						SUM(CASE WHEN a.order_ID IS NOT NULL AND b.order_ID IS  NULL THEN 1 ELSE 0 END) as order_ID,
						SUM(CASE WHEN a.market_side IS NOT NULL AND b.market_side IS  NULL THEN 1 ELSE 0 END) as market_side,
						SUM(CASE WHEN a.security IS NOT NULL AND b.security IS NULL THEN 1 ELSE 0 END) as security,
						SUM(CASE WHEN round(a.price,4) IS NOT NULL AND round(b.price,4) IS NULL THEN 1 ELSE 0 END) as price,
						SUM(CASE WHEN round(a.quantity,4) IS NOT NULL AND round(b.quantity,4) IS NULL THEN 1 ELSE 0 END) as quantity,
						SUM(CASE WHEN a.undisclosed_vol IS NOT NULL AND b.undisclosed_vol IS  NULL THEN 1 ELSE 0 END) as undisclosed_vol,
						SUM(CASE WHEN a.value1 IS NOT NULL AND b.value1 IS  NULL THEN 1 ELSE 0 END) as value1,
						SUM(CASE WHEN trim(a.order_flags) IS NOT NULL AND trim(b.order_flags) IS  NULL THEN 1 ELSE 0 END) as order_flags,
						SUM(CASE WHEN a.account IS NOT NULL AND b.account IS  NULL THEN 1 ELSE 0 END) as account,
						SUM(CASE WHEN a.accountType IS NOT NULL AND b.accountType IS  NULL THEN 1 ELSE 0 END) as accountType,
						SUM(CASE WHEN a.broker IS NOT NULL AND b.broker IS  NULL THEN 1 ELSE 0 END) as broker,
						SUM(CASE WHEN a.trader IS NOT NULL AND b.trader IS  NULL THEN 1 ELSE 0 END) as trader,
						SUM(CASE WHEN a.order_type IS NOT NULL AND b.order_type IS  NULL THEN 1 ELSE 0 END) as order_type,
						SUM(CASE WHEN a.parent_order_id IS NOT NULL AND b.parent_order_id IS  NULL THEN 1 ELSE 0 END) as parent_order_id,
						SUM(CASE WHEN a.execution_instrcutions IS NOT NULL AND b.execution_instrcutions IS  NULL THEN 1 ELSE 0 END) as execution_instrcutions,
						SUM(CASE WHEN a.order_receive_date IS NOT NULL AND b.order_receive_date IS  NULL THEN 1 ELSE 0 END) as order_receive_date,
						SUM(CASE WHEN a.order_receive_time IS NOT NULL AND b.order_receive_time IS  NULL THEN 1 ELSE 0 END) as order_receive_time,
						SUM(CASE WHEN a.other IS NOT NULL AND b.other IS  NULL THEN 1 ELSE 0 END) as other

						FROM LC.SMARTS_ENTER_QA_''' + region + '_' + busDt + ''' a
						INNER JOIN LC.SMARTS_ENTER_PROD_''' + region + '_' + busDt + ''' b
						ON a.order_ID = b.order_ID
					 
						UNION ALL
						
						SELECT
						'QA = NULL AND PROD = NULL' AS compareTpye,
						SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS  NULL THEN 1 ELSE 0 END) as exchange_id,
						SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS  NULL THEN 1 ELSE 0 END) as message_type,
						SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS  NULL THEN 1 ELSE 0 END) as date1,
						SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':')  IS  NULL THEN 1 ELSE 0 END) as time1,
						SUM(CASE WHEN a.order_ID IS NULL AND b.order_ID IS  NULL THEN 1 ELSE 0 END) as order_ID,
						SUM(CASE WHEN a.market_side IS NULL AND b.market_side IS  NULL THEN 1 ELSE 0 END) as market_side,
						SUM(CASE WHEN a.security IS NULL AND b.security IS NULL THEN 1 ELSE 0 END) as security,
						SUM(CASE WHEN round(a.price,4) IS NULL AND round(b.price,4) IS NULL THEN 1 ELSE 0 END) as price,
						SUM(CASE WHEN round(a.quantity,4) IS NULL AND round(b.quantity,4) IS NULL THEN 1 ELSE 0 END) as quantity,
						SUM(CASE WHEN a.undisclosed_vol IS NULL AND b.undisclosed_vol IS  NULL THEN 1 ELSE 0 END) as undisclosed_vol,
						SUM(CASE WHEN a.value1 IS NULL AND b.value1 IS  NULL THEN 1 ELSE 0 END) as value1,
						SUM(CASE WHEN trim(a.order_flags) IS NULL AND trim(b.order_flags) IS  NULL THEN 1 ELSE 0 END) as order_flags,
						SUM(CASE WHEN a.account IS NULL AND b.account IS  NULL THEN 1 ELSE 0 END) as account,
						SUM(CASE WHEN a.accountType IS NULL AND b.accountType IS  NULL THEN 1 ELSE 0 END) as accountType,
						SUM(CASE WHEN a.broker IS NULL AND b.broker IS  NULL THEN 1 ELSE 0 END) as broker,
						SUM(CASE WHEN a.trader IS NULL AND b.trader IS  NULL THEN 1 ELSE 0 END) as trader,
						SUM(CASE WHEN a.order_type IS NULL AND b.order_type IS  NULL THEN 1 ELSE 0 END) as order_type,
						SUM(CASE WHEN a.parent_order_id IS NULL AND b.parent_order_id IS  NULL THEN 1 ELSE 0 END) as parent_order_id,
						SUM(CASE WHEN a.execution_instrcutions IS NULL AND b.execution_instrcutions IS  NULL THEN 1 ELSE 0 END) as execution_instrcutions,
						SUM(CASE WHEN a.order_receive_date IS NULL AND b.order_receive_date IS  NULL THEN 1 ELSE 0 END) as order_receive_date,
						SUM(CASE WHEN a.order_receive_time IS NULL AND b.order_receive_time IS  NULL THEN 1 ELSE 0 END) as order_receive_time,
						SUM(CASE WHEN a.other IS NULL AND b.other IS  NULL THEN 1 ELSE 0 END) as other

						FROM LC.SMARTS_ENTER_QA_''' + region + '_' + busDt + ''' a
						INNER JOIN LC.SMARTS_ENTER_PROD_''' + region + '_' + busDt + ''' b
						ON a.order_ID = b.order_ID
						
						UNION ALL
						
						SELECT
						'QA = PROD' AS compareType,
						SUM(CASE WHEN a.exchange_id = b.exchange_id THEN 1 ELSE 0 END) as exchange_id,
						SUM(CASE WHEN a.message_type = b.message_type THEN 1 ELSE 0 END) as message_type,
						SUM(CASE WHEN a.date1 = b.date1 THEN 1 ELSE 0 END) as date1,
						SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') = replace(substr(b.time1,1,8),'.',':')  THEN 1 ELSE 0 END) as time1,
						SUM(CASE WHEN a.order_ID = b.order_ID THEN 1 ELSE 0 END) as order_ID,
						SUM(CASE WHEN a.market_side = b.market_side THEN 1 ELSE 0 END) as market_side,
						SUM(CASE WHEN a.security = b.security THEN 1 ELSE 0 END) as security,
						SUM(CASE WHEN round(a.price,4) = round(b.price,4) THEN 1 ELSE 0 END) as price,
						SUM(CASE WHEN round(a.quantity,4) = round(b.quantity,4) THEN 1 ELSE 0 END) as quantity,
						SUM(CASE WHEN a.undisclosed_vol = b.undisclosed_vol THEN 1 ELSE 0 END) as undisclosed_vol,
						SUM(CASE WHEN a.value1 = b.value1 THEN 1 ELSE 0 END) as value1,
						SUM(CASE WHEN trim(a.order_flags) = trim(b.order_flags) THEN 1 ELSE 0 END) as order_flags,
						SUM(CASE WHEN a.account = b.account THEN 1 ELSE 0 END) as account,
						SUM(CASE WHEN a.accountType = b.accountType THEN 1 ELSE 0 END) as accountType,
						SUM(CASE WHEN a.broker = b.broker THEN 1 ELSE 0 END) as broker,
						SUM(CASE WHEN a.trader = b.trader THEN 1 ELSE 0 END) as trader,
						SUM(CASE WHEN a.order_type = b.order_type THEN 1 ELSE 0 END) as order_type,
						SUM(CASE WHEN a.parent_order_id = b.parent_order_id THEN 1 ELSE 0 END) as parent_order_id,
						SUM(CASE WHEN a.execution_instrcutions = b.execution_instrcutions THEN 1 ELSE 0 END) as execution_instrcutions,
						SUM(CASE WHEN a.order_receive_date = b.order_receive_date THEN 1 ELSE 0 END) as order_receive_date,
						SUM(CASE WHEN a.order_receive_time = b.order_receive_time THEN 1 ELSE 0 END) as order_receive_time,
						SUM(CASE WHEN a.other = b.other THEN 1 ELSE 0 END) as other
						
						FROM LC.SMARTS_ENTER_QA_''' + region + '_' + busDt + ''' a
						INNER JOIN LC.SMARTS_ENTER_PROD_''' + region + '_' + busDt + ''' b
						ON a.order_ID = b.order_ID
						
						UNION ALL
						
						SELECT
						'MATCH Count' AS compareType,
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID),
						COUNT(a.order_ID)
						FROM LC.SMARTS_ENTER_QA_''' + region + '_' + busDt + ''' a
						INNER JOIN LC.SMARTS_ENTER_PROD_''' + region + '_' + busDt + ''' b
						ON a.order_ID = b.order_ID
						'''
					 )
		
		## Fetching the Table columns from Database
		num_fields = len(dbcur.description)
		fields = [i[0] for i in dbcur.description]
		
		## Writing the above query output to database
		with open('enterOutput.csv', 'w') as result:
			writer = csv.writer(result, dialect = 'excel')
			writer.writerow(fields)
			for records in dbcur.fetchall():
				writer.writerow(records)
		
		## Creating the DF from csv
		enterDF = pd.DataFrame()
		enterDF = pd.read_csv('enterOutput.csv')
		
		#Transposing the DF
		self.enterDFTr = enterDF.transpose()
		
		print ('       Enter validation completed')
	
	## This function will verify the AMEND table
	def verifyAmendTable(self, region, busDt):
		dbcur.execute(
				  '''
					SELECT
					'QA = PROD' AS compareType,
					SUM(CASE WHEN a.exchange_id = b.exchange_id THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type = b.message_type THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 = b.date1 THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':')= replace(substr(a.time1,1,8),'.',':') THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID = b.order_ID THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.new_order_ID = b.new_order_ID THEN 1 ELSE 0 END) as new_order_ID,
					SUM(CASE WHEN a.market_side = b.market_side THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security = b.security THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN round(a.price,4) = round(b.price,4) THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN a.quantity = b.quantity THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.undisclosed_vol = b.undisclosed_vol THEN 1 ELSE 0 END) as undisclosed_vol,
					SUM(CASE WHEN a.value1 = b.value1 THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN trim(a.order_flags) = trim(b.order_flags) THEN 1 ELSE 0 END) as order_flags,
					SUM(CASE WHEN a.order_type = b.order_type THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id = b.parent_order_id THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.execution_instrcutions = b.execution_instrcutions THEN 1 ELSE 0 END) as execution_instrcutions,
					SUM(CASE WHEN a.amend_receive_date = b.amend_receive_date THEN 1 ELSE 0 END) as amend_receive_date,
					SUM(CASE WHEN  replace(substr(a.amend_receive_time,1,8),'.',':')= replace(substr(b.amend_receive_time,1,8),'.',':') THEN 1 ELSE 0 END) as amend_receive_time,
					SUM(CASE WHEN a.other = b.other THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_AMEND_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_AMEND_PROD_''' + region + '_' + busDt + ''' b
					ON a.new_order_ID = b.new_order_ID

					UNION ALL

					SELECT
					'QA = NULL AND PROD = NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID IS NULL AND b.order_ID IS NULL THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.new_order_ID IS NULL AND b.new_order_ID IS NULL THEN 1 ELSE 0 END) as new_order_ID,
					SUM(CASE WHEN a.market_side IS NULL AND b.market_side IS NULL THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security IS NULL AND b.security IS NULL THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN round(a.price,4) IS NULL AND round(b.price,4) IS NULL THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN a.quantity IS NULL AND b.quantity IS NULL THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.undisclosed_vol IS NULL AND b.undisclosed_vol IS NULL THEN 1 ELSE 0 END) as undisclosed_vol,
					SUM(CASE WHEN a.value1 IS NULL AND b.value1 IS NULL THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.order_flags IS NULL AND b.order_flags IS NULL THEN 1 ELSE 0 END) as order_flags,
					SUM(CASE WHEN a.order_type IS NULL AND b.order_type IS NULL THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id IS NULL AND b.parent_order_id IS NULL THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.execution_instrcutions IS NULL AND b.execution_instrcutions IS NULL THEN 1 ELSE 0 END) as execution_instrcutions,
					SUM(CASE WHEN a.amend_receive_date IS NULL AND b.amend_receive_date IS NULL THEN 1 ELSE 0 END) as amend_receive_date,
					SUM(CASE WHEN a.amend_receive_time IS NULL AND b.amend_receive_time IS NULL THEN 1 ELSE 0 END) as amend_receive_time,
					SUM(CASE WHEN a.other IS NULL AND b.other IS NULL THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_AMEND_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_AMEND_PROD_''' + region + '_' + busDt + ''' b
					ON a.new_order_ID = b.new_order_ID

					UNION ALL

					SELECT
					'QA NOT NULL AND PROD = NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NOT NULL AND b.exchange_id IS NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NOT NULL AND b.message_type IS NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NOT NULL AND b.date1 IS NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NOT NULL AND replace(substr(b.time1,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID IS NOT NULL AND b.order_ID IS NULL THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.new_order_ID IS NOT NULL AND b.new_order_ID IS NULL THEN 1 ELSE 0 END) as new_order_ID,
					SUM(CASE WHEN a.market_side IS NOT NULL AND b.market_side IS NULL THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security IS NOT NULL AND b.security IS NULL THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN round(a.price,4) IS NOT NULL AND round(b.price,4) IS NULL THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN a.quantity IS NOT NULL AND b.quantity IS NULL THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.undisclosed_vol IS NOT NULL AND b.undisclosed_vol IS NULL THEN 1 ELSE 0 END) as undisclosed_vol,
					SUM(CASE WHEN a.value1 IS NOT NULL AND b.value1 IS NULL THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.order_flags IS NOT NULL AND b.order_flags IS NULL THEN 1 ELSE 0 END) as order_flags,
					SUM(CASE WHEN a.order_type IS NOT NULL AND b.order_type IS NULL THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id IS NOT NULL AND b.parent_order_id IS NULL THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.execution_instrcutions IS NOT NULL AND b.execution_instrcutions IS NULL THEN 1 ELSE 0 END) as execution_instrcutions,
					SUM(CASE WHEN a.amend_receive_date IS NOT NULL AND b.amend_receive_date IS NULL THEN 1 ELSE 0 END) as amend_receive_date,
					SUM(CASE WHEN a.amend_receive_time IS NOT NULL AND b.amend_receive_time IS NULL THEN 1 ELSE 0 END) as amend_receive_time,
					SUM(CASE WHEN a.other IS NOT NULL AND b.other IS NULL THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_AMEND_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_AMEND_PROD_''' + region + '_' + busDt + ''' b
					ON a.new_order_ID = b.new_order_ID

					UNION ALL

					SELECT
					'QA = NULL AND PROD NOT NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NOT NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NOT NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NOT NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NOT NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID IS NULL AND b.order_ID IS NOT NULL THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.new_order_ID IS NULL AND b.new_order_ID IS NOT NULL THEN 1 ELSE 0 END) as new_order_ID,
					SUM(CASE WHEN a.market_side IS NULL AND b.market_side IS NOT NULL THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security IS NULL AND b.security IS NOT NULL THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN round(a.price,4) IS NULL AND round(b.price,4) IS NOT NULL THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN a.quantity IS NULL AND b.quantity IS NOT NULL THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.undisclosed_vol IS NULL AND b.undisclosed_vol IS NOT NULL THEN 1 ELSE 0 END) as undisclosed_vol,
					SUM(CASE WHEN a.value1 IS NULL AND b.value1 IS NOT NULL THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.order_flags IS NULL AND b.order_flags IS NOT NULL THEN 1 ELSE 0 END) as order_flags,
					SUM(CASE WHEN a.order_type IS NULL AND b.order_type IS NOT NULL THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id IS NULL AND b.parent_order_id IS NOT NULL THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.execution_instrcutions IS NULL AND b.execution_instrcutions IS NOT NULL THEN 1 ELSE 0 END) as execution_instrcutions,
					SUM(CASE WHEN a.amend_receive_date IS NULL AND b.amend_receive_date IS NOT NULL THEN 1 ELSE 0 END) as amend_receive_date,
					SUM(CASE WHEN a.amend_receive_time IS NULL AND b.amend_receive_time IS NOT NULL THEN 1 ELSE 0 END) as amend_receive_time,
					SUM(CASE WHEN a.other IS NULL AND b.other IS NOT NULL THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_AMEND_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_AMEND_PROD_''' + region + '_' + busDt + ''' b
					ON a.new_order_ID = b.new_order_ID
					
					UNION ALL
						
					SELECT
					'MATCH Count' AS compareType,
					COUNT(a.new_order_ID),
					COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
                    COUNT(a.new_order_ID),
					COUNT(a.new_order_ID),
					COUNT(a.new_order_ID),
					COUNT(a.new_order_ID)
					FROM LC.SMARTS_AMEND_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_AMEND_PROD_''' + region + '_' + busDt + ''' b
					ON a.new_order_ID = b.new_order_ID
					
					'''
				)
		
		## Fetching the Table columns from Database
		num_fields = len(dbcur.description)
		fields = [i[0] for i in dbcur.description]
		
		## Writing the above query output to database			 
		with open('amendOutput.csv', 'w') as result:
			writer = csv.writer(result, dialect = 'excel')
			writer.writerow(fields)
			for records in dbcur.fetchall():
				writer.writerow(records)
		
		## Creating the DF from csv
		amendDF = pd.DataFrame()
		amendDF = pd.read_csv('amendOutput.csv')
		self.amendDFTr = amendDF.transpose()
		print ('       Amend validation completed')
	
	## This function will verify the DELET table
	def verifyDeletTable(self, region, busDt):
		dbcur.execute(
				  '''
					SELECT
					'QA = PROD' AS compareType,
					SUM(CASE WHEN a.exchange_id = b.exchange_id THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type = b.message_type THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 = b.date1 THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') = replace(substr(b.time1,1,8),'.',':') THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID = b.order_ID THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.market_side = b.market_side THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security = b.security THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN a.order_type = b.order_type THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id = b.parent_order_id THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.delete_receive_date = b.delete_receive_date THEN 1 ELSE 0 END) as delete_receive_date,
					SUM(CASE WHEN replace(substr(a.delete_receive_time,1,8),'.',':') = replace(substr(b.delete_receive_time,1,8),'.',':') THEN 1 ELSE 0 END) as delete_receive_time

					FROM LC.SMARTS_DELET_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_DELET_PROD_''' + region + '_' + busDt + ''' b
					ON a.order_ID = b.order_ID

					UNION ALL

					SELECT
					'QA = NULL AND PROD NOT NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID IS NULL AND b.order_ID IS NULL THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.market_side IS NULL AND b.market_side IS NULL THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security IS NULL AND b.security IS NULL THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN a.order_type IS NULL AND b.order_type IS NULL THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id IS NULL AND b.parent_order_id IS NULL THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.delete_receive_date IS NULL AND b.delete_receive_date IS NULL THEN 1 ELSE 0 END) as delete_receive_date,
					SUM(CASE WHEN replace(substr(a.delete_receive_time,1,8),'.',':') IS NULL AND replace(substr(b.delete_receive_time,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as delete_receive_time

					FROM LC.SMARTS_DELET_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_DELET_PROD_''' + region + '_' + busDt + ''' b
					ON a.order_ID = b.order_ID

					UNION ALL

					SELECT
					'QA NOT NULL AND PROD = NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NOT NULL AND b.exchange_id IS NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NOT NULL AND b.message_type IS NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NOT NULL AND b.date1 IS NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NOT NULL AND replace(substr(b.time1,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID IS NOT NULL AND b.order_ID IS NULL THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.market_side IS NOT NULL AND b.market_side IS NULL THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security IS NOT NULL AND b.security IS NULL THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN a.order_type IS NOT NULL AND b.order_type IS NULL THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id IS NOT NULL AND b.parent_order_id IS NULL THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.delete_receive_date IS NOT NULL AND b.delete_receive_date IS NULL THEN 1 ELSE 0 END) as delete_receive_date,
					SUM(CASE WHEN replace(substr(a.delete_receive_time,1,8),'.',':') IS NOT NULL AND replace(substr(b.delete_receive_time,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as delete_receive_time

					FROM LC.SMARTS_DELET_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_DELET_PROD_''' + region + '_' + busDt + ''' b
					ON a.order_ID = b.order_ID

					UNION ALL 

					SELECT
					'QA IS NULL PROD IS NOT NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NOT NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NOT NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NOT NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NOT NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.order_ID IS NULL AND b.order_ID IS NOT NULL THEN 1 ELSE 0 END) as order_ID,
					SUM(CASE WHEN a.market_side IS NULL AND b.market_side IS NOT NULL THEN 1 ELSE 0 END) as market_side,
					SUM(CASE WHEN a.security IS NULL AND b.security IS NOT NULL THEN 1 ELSE 0 END) as security,
					SUM(CASE WHEN a.order_type IS NULL AND b.order_type IS NOT NULL THEN 1 ELSE 0 END) as order_type,
					SUM(CASE WHEN a.parent_order_id IS NULL AND b.parent_order_id IS NOT NULL THEN 1 ELSE 0 END) as parent_order_id,
					SUM(CASE WHEN a.delete_receive_date IS NULL AND b.delete_receive_date IS NOT NULL THEN 1 ELSE 0 END) as delete_receive_date,
					SUM(CASE WHEN replace(substr(a.delete_receive_time,1,8),'.',':') IS NULL AND replace(substr(b.delete_receive_time,1,8),'.',':') IS NOT NULL THEN 1 ELSE 0 END) as delete_receive_time

					FROM LC.SMARTS_DELET_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_DELET_PROD_''' + region + '_' + busDt + ''' b
					ON a.order_ID = b.order_ID
					
					UNION ALL
						
					SELECT
					'MATCH Count' AS compareType,
					COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID),
                    COUNT(a.order_ID)
					FROM LC.SMARTS_DELET_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_DELET_PROD_''' + region + '_' + busDt + ''' b
					ON a.order_ID = b.order_ID
					
					'''
				)
		
		## Fetching the Table columns from Database
		num_fields = len(dbcur.description)
		fields = [i[0] for i in dbcur.description]
			
		## Writing the above query output to database		
		with open('deletOutput.csv', 'w') as result:
			writer = csv.writer(result, dialect = 'excel')
			writer.writerow(fields)
			for records in dbcur.fetchall():
				writer.writerow(records)
		
		## Creating the DF from csv
		#writer = ExcelWriter('Output.xlsx')
		deletDF = pd.DataFrame()
		deletDF = pd.read_csv('deletOutput.csv')
		self.deletDFTr = deletDF.transpose()
		print ('       Delet validation completed')
	
	## This function will verify the TRADE table
	def verifyTradeTable(self, region, busDt):
		dbcur.execute(
				  '''
					SELECT
					'QA = PROD' AS compareType,
					SUM(CASE WHEN a.exchange_id = b.exchange_id THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type = b.message_type THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 = b.date1 THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') = replace(substr(b.time1,1,8),'.',':') THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.trade_ID = b.trade_ID THEN 1 ELSE 0 END) as trade_ID,
					--SUM(CASE WHEN locate(B.SECURITY,(case when  c.ISIN is null then a.SECURITY else  c.ISIN end ))=1 then 1 else 0 end) as security,
					SUM(CASE WHEN a.SECURITY = b.SECURITY THEN 1 ELSE 0 END) as Security,
					SUM(CASE WHEN round(a.price,4) = round(b.price,4) THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN round(a.quantity,4) = round(b.quantity,4) THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.trade_flags = b.trade_flags THEN 1 ELSE 0 END) as trade_flags,
					SUM(CASE WHEN a.value1 = b.value1 THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.bid_order_id = b.bid_order_id THEN 1 ELSE 0 END) as bid_order_id,
					SUM(CASE WHEN a.ask_order_id = b.ask_order_id THEN 1 ELSE 0 END) as ask_order_id,
					SUM(CASE WHEN a.trade_type = b.trade_type THEN 1 ELSE 0 END) as trade_type,
					SUM(CASE WHEN a.parent_trade_id = b.parent_trade_id THEN 1 ELSE 0 END) as parent_trade_id,
					SUM(CASE WHEN a.other = b.other THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_TRADE_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_TRADE_PROD_''' + region + '_' + busDt + ''' b
					ON a.trade_ID = b.trade_ID

					UNION ALL

					SELECT
					'QA = NULL AND PROD = NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.trade_ID IS NULL AND b.trade_ID IS NULL THEN 1 ELSE 0 END) as trade_ID,
					--SUM(CASE WHEN (case when  c.ISIN is null then B.SECURITY else  c.ISIN end ) IS NULL AND B.SECURITY IS NULL then 1 else 0 end) as  security,
					SUM(CASE WHEN a.SECURITY IS NULL AND b.SECURITY IS NULL THEN 1 ELSE 0 END) as Securtiy,
					SUM(CASE WHEN round(a.price,4) IS NULL AND round(b.price,4) IS NULL THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN round(a.quantity,4) IS NULL AND round(b.quantity,4) IS NULL THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.trade_flags IS NULL AND b.trade_flags IS NULL THEN 1 ELSE 0 END) as trade_flags,
					SUM(CASE WHEN a.value1 IS NULL AND b.value1 IS NULL THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.bid_order_id IS NULL AND b.bid_order_id IS NULL THEN 1 ELSE 0 END) as bid_order_id,
					SUM(CASE WHEN a.ask_order_id IS NULL AND b.ask_order_id IS NULL THEN 1 ELSE 0 END) as ask_order_id,
					SUM(CASE WHEN a.trade_type IS NULL AND b.trade_type IS NULL THEN 1 ELSE 0 END) as trade_type,
					SUM(CASE WHEN a.parent_trade_id IS NULL AND b.parent_trade_id IS NULL THEN 1 ELSE 0 END) as parent_trade_id,
					SUM(CASE WHEN a.other IS NULL AND b.other IS NULL THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_TRADE_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_TRADE_PROD_''' + region + '_' + busDt + ''' b
					ON a.trade_ID = b.trade_ID

					UNION ALL

					SELECT
					'QA NOT NULL AND PROD = NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NOT NULL AND b.exchange_id IS NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NOT NULL AND b.message_type IS NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NOT NULL AND b.date1 IS NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NOT NULL AND replace(substr(b.time1,1,8),'.',':') IS NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.trade_ID IS NOT NULL AND b.trade_ID IS NULL THEN 1 ELSE 0 END) as trade_ID,
					--SUM(CASE WHEN (case when  c.ISIN is null then B.SECURITY else  c.ISIN end )  IS NOT NULL AND B.SECURITY IS NULL then 1 else 0 end) as security,
					SUM(CASE WHEN a.SECURITY IS NOT NULL AND b.SECURITY IS NULL THEN 1 ELSE 0 END) as Securtiy,
					SUM(CASE WHEN round(a.price,4) IS NOT NULL AND round(b.price,4) IS NULL THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN round(a.quantity,4) IS NOT NULL AND round(b.quantity,4) IS NULL THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.trade_flags IS NOT NULL AND b.trade_flags IS NULL THEN 1 ELSE 0 END) as trade_flags,
					SUM(CASE WHEN a.value1 IS NOT NULL AND b.value1 IS NULL THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.bid_order_id IS NOT NULL AND b.bid_order_id IS NULL THEN 1 ELSE 0 END) as bid_order_id,
					SUM(CASE WHEN a.ask_order_id IS NOT NULL AND b.ask_order_id IS NULL THEN 1 ELSE 0 END) as ask_order_id,
					SUM(CASE WHEN a.trade_type IS NOT NULL AND b.trade_type IS NULL THEN 1 ELSE 0 END) as trade_type,
					SUM(CASE WHEN a.parent_trade_id IS NOT NULL AND b.parent_trade_id IS NULL THEN 1 ELSE 0 END) as parent_trade_id,
					SUM(CASE WHEN a.other IS NOT NULL AND b.other IS NULL THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_TRADE_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_TRADE_PROD_''' + region + '_' + busDt + ''' b
					ON a.trade_ID = b.trade_ID


					UNION ALL

					SELECT
					'QA = NULL AND PROD NOT NULL' AS compareType,
					SUM(CASE WHEN a.exchange_id IS NULL AND b.exchange_id IS NOT NULL THEN 1 ELSE 0 END) as exchange_id,
					SUM(CASE WHEN a.message_type IS NULL AND b.message_type IS NOT NULL THEN 1 ELSE 0 END) as message_type,
					SUM(CASE WHEN a.date1 IS NULL AND b.date1 IS NOT NULL THEN 1 ELSE 0 END) as date1,
					SUM(CASE WHEN replace(substr(a.time1,1,8),'.',':') IS NULL AND replace(substr(b.time1,1,8),'.',':') IS NOT NULL THEN 1 ELSE 0 END) as time1,
					SUM(CASE WHEN a.trade_ID IS NULL AND b.trade_ID IS NOT NULL THEN 1 ELSE 0 END) as trade_ID,
					--SUM(CASE WHEN (case when  c.ISIN is null then B.SECURITY else  c.ISIN end ) IS NULL AND B.SECURITY IS NOT NULL then 1 else 0 end) as  security,
					SUM(CASE WHEN a.SECURITY IS NULL AND b.SECURITY IS NOT NULL THEN 1 ELSE 0 END) as Securtiy,
					SUM(CASE WHEN round(a.price,4) IS NULL AND round(b.price,4) IS NOT NULL THEN 1 ELSE 0 END) as price,
					SUM(CASE WHEN round(a.quantity,4) IS NULL AND round(b.quantity,4) IS NOT NULL THEN 1 ELSE 0 END) as quantity,
					SUM(CASE WHEN a.trade_flags IS NULL AND b.trade_flags IS NOT NULL THEN 1 ELSE 0 END) as trade_flags,
					SUM(CASE WHEN a.value1 IS NULL AND b.value1 IS NOT NULL THEN 1 ELSE 0 END) as value1,
					SUM(CASE WHEN a.bid_order_id IS NULL AND b.bid_order_id IS NOT NULL THEN 1 ELSE 0 END) as bid_order_id,
					SUM(CASE WHEN a.ask_order_id IS NULL AND b.ask_order_id IS NOT NULL THEN 1 ELSE 0 END) as ask_order_id,
					SUM(CASE WHEN a.trade_type IS NULL AND b.trade_type IS NOT NULL THEN 1 ELSE 0 END) as trade_type,
					SUM(CASE WHEN a.parent_trade_id IS NULL AND b.parent_trade_id IS NOT NULL THEN 1 ELSE 0 END) as parent_trade_id,
					SUM(CASE WHEN a.other IS NULL AND b.other IS NOT NULL THEN 1 ELSE 0 END) as other

					FROM LC.SMARTS_TRADE_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_TRADE_PROD_''' + region + '_' + busDt + ''' b
					ON a.trade_ID = b.trade_ID
					
					UNION ALL
						
					SELECT
					'MATCH Count' AS compareType,
					COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID),
                    COUNT(a.trade_ID)
					FROM LC.SMARTS_TRADE_QA_''' + region + '_' + busDt + ''' a
					INNER JOIN LC.SMARTS_TRADE_PROD_''' + region + '_' + busDt + ''' b
					ON a.trade_ID = b.trade_ID
					
					'''
				)
		
		## Fetching the Table columns from Database
		num_fields = len(dbcur.description)
		fields = [i[0] for i in dbcur.description]
		
		## Fetching the Table columns from Database
		with open('tradeOutput.csv', 'w') as result:
			writer = csv.writer(result, dialect = 'excel')
			writer.writerow(fields)
			for records in dbcur.fetchall():
				writer.writerow(records)
		
		## Writing the above query output to database
		tradeDF = pd.DataFrame()
		tradeDF = pd.read_csv('tradeOutput.csv')
		self.tradeDFTr = tradeDF.transpose()
		print ('       Trade validation completed')
		
	## This function will write the transposed values to Excel
	def generateReport(self, region, busDt):
		print('Generating final report...')
		xlWriter = ExcelWriter('SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
		self.enterDFTr.to_excel(xlWriter, 'Enter')
		self.amendDFTr.to_excel(xlWriter, 'Amend')
		self.deletDFTr.to_excel(xlWriter, 'Delet')
		self.tradeDFTr.to_excel(xlWriter, 'Trade')
		
		time.sleep(5)
		xlWriter.save()
		time.sleep(5)
		xlWriter.close()
		
	## Deleting temporary .csv files
		print('Removing temporary files...')
		dir = os.getcwd()
		folders = os.listdir(dir)
		
		if 'enterOutput.csv' in folders:
			os.remove('enterOutput.csv')
			
		if 'amendOutput.csv' in folders:
			os.remove('amendOutput.csv')
			
		if 'deletOutput.csv' in folders:
			os.remove('deletOutput.csv')
			
		if 'tradeOutput.csv' in folders:
			os.remove('tradeOutput.csv')
		
		if 'cantrOutput.csv' in folders:
			os.remove('cantrOutput.csv')
		
		if 'offtrOutput.csv' in folders:
			os.remove('offtrOutput.csv')
			
	## This function will Sum up the total for all 4 columns in Excel
	def sumUpValues(self, region, busDt):
		dir = os.getcwd()
		
	## Summing up for Sheet ENTER
		wb = Workbook()
		wb = load_workbook(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
		ws = wb.get_sheet_by_name('Enter')
	
		i = 1
		while(i < (len(ws.rows))-1):
			ws["G2"].value = 'Difference'
			ws["G3"].value = "=F3-SUM(B3:E3)"
			ws['G4'].value = "=F4-SUM(B4:E4)"
			ws['G5'].value = "=F5-SUM(B5:E5)"
			ws['G6'].value = "=F6-SUM(B6:E6)"
			ws['G7'].value = "=F7-SUM(B7:E7)"
			ws['G8'].value = "=F8-SUM(B8:E8)"
			ws['G9'].value = "=F9-SUM(B9:E9)"
			ws['G10'].value = "=F10-SUM(B10:E10)"
			ws['G11'].value = "=F11-SUM(B11:E11)"
			ws['G12'].value = "=F12-SUM(B12:E12)"
			ws['G13'].value = "=F13-SUM(B13:E13)"
			ws['G14'].value = "=F14-SUM(B14:E14)"
			ws['G15'].value = "=F15-SUM(B15:E15)"
			ws['G16'].value = "=F16-SUM(B16:E16)"
			ws['G17'].value = "=F17-SUM(B17:E17)"
			ws['G18'].value = "=F18-SUM(B18:E18)"
			ws['G19'].value = "=F19-SUM(B19:E19)"
			ws['G20'].value = "=F20-SUM(B20:E20)"
			ws['G21'].value = "=F21-SUM(B21:E21)"
			ws['G22'].value = "=F22-SUM(B22:E22)"
			ws['G23'].value = "=F23-SUM(B23:E23)"
			ws['G24'].value = "=F24-SUM(B24:E24)"
			i+=1
			
		#wb.save(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
		
	## Summing up for sheet Amend
		ws = wb.get_sheet_by_name('Amend')
		i = 1
		while(i < (len(ws.rows))-1):
			ws["G2"].value = 'Difference'
			ws["G3"].value = "=F3-SUM(B3:E3)"
			ws['G4'].value = "=F4-SUM(B4:E4)"
			ws['G5'].value = "=F5-SUM(B5:E5)"
			ws['G6'].value = "=F6-SUM(B6:E6)"
			ws['G7'].value = "=F7-SUM(B7:E7)"
			ws['G8'].value = "=F8-SUM(B8:E8)"
			ws['G9'].value = "=F9-SUM(B9:E9)"
			ws['G10'].value = "=F10-SUM(B10:E10)"
			ws['G11'].value = "=F11-SUM(B11:E11)"
			ws['G12'].value = "=F12-SUM(B12:E12)"
			ws['G13'].value = "=F13-SUM(B13:E13)"
			ws['G14'].value = "=F14-SUM(B14:E14)"
			ws['G15'].value = "=F15-SUM(B15:E15)"
			ws['G16'].value = "=F16-SUM(B16:E16)"
			ws['G17'].value = "=F17-SUM(B17:E17)"
			ws['G18'].value = "=F18-SUM(B18:E18)"
			ws['G19'].value = "=F19-SUM(B19:E19)"
			ws['G20'].value = "=F20-SUM(B20:E20)"
			ws['G21'].value = "=F21-SUM(B21:E21)"
			i+=1
			
		#wb.save(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
		
	## Summing up for sheet Delet
		ws = wb.get_sheet_by_name('Delet')
		i = 1
		while(i < (len(ws.rows))-1):
			ws["G2"].value = 'Difference'
			ws["G3"].value = "=F3-SUM(B3:E3)"
			ws['G4'].value = "=F4-SUM(B4:E4)"
			ws['G5'].value = "=F5-SUM(B5:E5)"
			ws['G6'].value = "=F6-SUM(B6:E6)"
			ws['G7'].value = "=F7-SUM(B7:E7)"
			ws['G8'].value = "=F8-SUM(B8:E8)"
			ws['G9'].value = "=F9-SUM(B9:E9)"
			ws['G10'].value = "=F10-SUM(B10:E10)"
			ws['G11'].value = "=F11-SUM(B11:E11)"
			ws['G12'].value = "=F12-SUM(B12:E12)"
			ws['G13'].value = "=F13-SUM(B13:E13)"
			i+=1
			
		#wb.save(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
		
	## Summing up for sheet Trade
		ws = wb.get_sheet_by_name('Trade')
		i = 1
		while(i < (len(ws.rows))-1):
			ws["G2"].value = 'Difference'
			ws["G3"].value = "=F3-SUM(B3:E3)"
			ws['G4'].value = "=F4-SUM(B4:E4)"
			ws['G5'].value = "=F5-SUM(B5:E5)"
			ws['G6'].value = "=F6-SUM(B6:E6)"
			ws['G7'].value = "=F7-SUM(B7:E7)"
			ws['G8'].value = "=F8-SUM(B8:E8)"
			ws['G9'].value = "=F9-SUM(B9:E9)"
			ws['G10'].value = "=F10-SUM(B10:E10)"
			ws['G11'].value = "=F11-SUM(B11:E11)"
			ws['G12'].value = "=F12-SUM(B12:E12)"
			ws['G13'].value = "=F13-SUM(B13:E13)"
			ws['G14'].value = "=F14-SUM(B14:E14)"
			ws['G15'].value = "=F15-SUM(B15:E15)"
			ws['G16'].value = "=F16-SUM(B16:E16)"
			ws['G17'].value = "=F17-SUM(B17:E17)"
			i+=1
		
	## Saving the workbook after summing up the values
		wb.save(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
		
	## This function will find out the Non zero values from difference column & pull out the samples from Database
	def getMismatchSamples(self, region, busDt):
		dir = os.getcwd()
		filename = openpyxl.load_workbook(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx', data_only=True, use_iterators=False)
		lstSheetNames = filename.get_sheet_names()
		
		k = 0
		primaryKey = None

		while (k <= len(lstSheetNames)-1):
			sheet = filename.get_sheet_by_name(lstSheetNames[k])
			
			## Setting up the primary keys to run the sample query
			if (lstSheetNames[k] == 'Enter'):
				primaryKey = 'ORDER_ID'
			elif (lstSheetNames[k] == 'Amend'):
				primaryKey = 'NEW_ORDER_ID'
				primaryKey1 = 'ORDER_ID'
			elif (lstSheetNames[k] == 'Delet'):
				primaryKey = 'ORDER_ID'
			elif (lstSheetNames[k] == 'Trade'):
				primaryKey = 'TRADE_ID'
			
			i = 2
			for rw in range(len(sheet.rows) -1):
				if (i == len(sheet.rows)):
					break
				else:
					sheet.rows[i][6].value
					qaColumn = str(sheet.rows[i][0].value)
					prodColumn = str(sheet.rows[i][0].value)
					
					if(str(sheet.rows[i][6].value) <> '0'):
						
						## Printing the mismatch data for Enter & Delet sheet
						if (lstSheetNames[k] == 'Enter') or (lstSheetNames[k] == 'Delet'):
							dbcur.execute(
										'''SELECT DISTINCT a.''' + primaryKey + ''',
										a.''' + str(sheet.rows[i][0].value) + ' as QA_' + qaColumn + ''',
										b.''' + str(sheet.rows[i][0].value) + ' as PROD_' + prodColumn + '''
										FROM LC.SMARTS_''' + lstSheetNames[k] + '_QA_' + region + '_' + busDt + ''' a
										INNER JOIN LC.SMARTS_''' + lstSheetNames[k] + '_PROD_' + region + '_' + busDt + ''' b
										ON a.''' + primaryKey + ''' = b.''' + primaryKey + ''' 
										AND a.''' + str(sheet.rows[i][0].value) + ''' <> b.''' + str(sheet.rows[i][0].value) + ''' 
										FETCH FIRST 7 ROWS ONLY 
										'''
										)

							## printing the DB Headers
							num_fields = len(dbcur.description)
							fields = [b[0] for b in dbcur.description]
							
							## writing temp mismatch.csv file.
							with open('mismatch.csv', 'w') as result:
								writer = csv.writer(result, dialect = 'excel')
								writer.writerow(fields)
								for records in dbcur.fetchall():
									writer.writerow(records)

							mismatchDF = pd.DataFrame()
							mismatchDF = pd.read_csv('mismatch.csv')

							if (not mismatchDF.empty):
								print ('\n')
								print ('\t' + lstSheetNames[k])
								# print (mismatchDF)
								
								#printing the mismatches in PrettyTable
								fp = open(dir + '/mismatch.csv', 'r')
								file = from_csv(fp)
								fp.close()
								print(file)
								
							# xlWriter = ExcelWriter(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
							# mismatchDF.to_excel(xlWriter, 'Mismatches')
							# xlWriter.save()
							
							# usedRowCnt = len(filename.get_sheet_by_name(lstSheetNames[k]).rows) + 3				
							
							# book = load_workbook(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx')
							# writer = ExcelWriter(dir + '/SMARTSAutomation_' + region + '_' + busDt + '_Output.xlsx', engine='openpyxl') 
							# writer.book = book
							# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
							# mismatchDF.to_excel(writer, startrow=usedRowCnt)
							# writer.save()
						
						
						## Printing the mismatches for sheet Trade
						if (lstSheetNames[k] == 'Trade'):
							dbcur.execute(
										'''SELECT DISTINCT a.''' + primaryKey + ''',
										a.''' + str(sheet.rows[i][0].value) + ' as QA_' + qaColumn + ''',
										b.''' + str(sheet.rows[i][0].value) + ' as PROD_' + prodColumn + '''
										FROM LC.SMARTS_''' + lstSheetNames[k] + '_QA_' + region + '_' + busDt + ''' a
										INNER JOIN LC.SMARTS_''' + lstSheetNames[k] + '_PROD_' + region + '_' + busDt + ''' b
										ON a.''' + primaryKey + ''' = b.''' + primaryKey + ''' 
										AND a.''' + str(sheet.rows[i][0].value) + ''' <> b.''' + str(sheet.rows[i][0].value) + ''' 
										FETCH FIRST 7 ROWS ONLY 
										'''
										)
							
							## printing the DB Headers
							num_fields = len(dbcur.description)
							fields = [b[0] for b in dbcur.description]
							
							## writing temp mismatch.csv file.
							with open('mismatch.csv', 'w') as result:
								writer = csv.writer(result, dialect = 'excel')
								writer.writerow(fields)
								for records in dbcur.fetchall():
									writer.writerow(records)

							mismatchDF = pd.DataFrame()
							mismatchDF = pd.read_csv('mismatch.csv')

							if (not mismatchDF.empty):
								print ('\n')
								print ('\t' + lstSheetNames[k])
								#print (mismatchDF)
								
								## printing the mismatches in PrettyTable
								fp = open(dir + '/mismatch.csv', 'r')
								file = from_csv(fp)
								fp.close()
								print(file)

						
						## printing the mismatches for AMEND sheet
						if (lstSheetNames[k] == 'Amend'):
							dbcur.execute(
										'''SELECT DISTINCT a.''' + primaryKey + ''', a.''' + primaryKey1 + ''',
										a.''' + str(sheet.rows[i][0].value) + ' as QA_' + qaColumn + ''',
										b.''' + str(sheet.rows[i][0].value) + ' as PROD_' + prodColumn + '''
										FROM LC.SMARTS_''' + lstSheetNames[k] + '_QA_' + region + '_' + busDt + ''' a
										INNER JOIN LC.SMARTS_''' + lstSheetNames[k] + '_PROD_' + region + '_' + busDt + ''' b
										ON a.''' + primaryKey + ''' = b.''' + primaryKey + ''' 
										AND a.''' + primaryKey1 + ''' = b.''' + primaryKey1 + ''' 
										AND a.''' + str(sheet.rows[i][0].value) + ''' <> b.''' + str(sheet.rows[i][0].value) + ''' 
										FETCH FIRST 7 ROWS ONLY 
										'''
										)
							
							## printing the DB Headers
							num_fields = len(dbcur.description)
							fields = [b[0] for b in dbcur.description]
							
							## writing temp mismatch.csv file.
							with open('mismatch.csv', 'w') as result:
								writer = csv.writer(result, dialect = 'excel')
								writer.writerow(fields)
								for records in dbcur.fetchall():
									writer.writerow(records)

							mismatchDF = pd.DataFrame()
							mismatchDF = pd.read_csv('mismatch.csv')

							if (not mismatchDF.empty):
								print ('\n')
								print ('\t' + lstSheetNames[k])
								#print (mismatchDF)
								
								## printing the mismatches in PrettyTable
								fp = open(dir + '/mismatch.csv', 'r')
								file = from_csv(fp)
								fp.close()
								print(file)
								
				i+=1
			k+=1

		dir = os.getcwd()
		folders = os.listdir(dir)
		
		## Removing the mismatch.csv temp file
		if 'mismatch.csv' in folders:
			os.remove('mismatch.csv')
			print ('\n')
